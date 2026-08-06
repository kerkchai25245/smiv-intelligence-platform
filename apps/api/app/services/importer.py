import re
from hashlib import sha256
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_national_id, normalize_national_id
from app.models import ImportIssue, ImportJob, Patient, User
from app.services.audit import add_audit
from app.services.patients import make_version, parse_bool, parse_date

ALIASES = {
    "national_id": {
        "national_id",
        "national id",
        "thai_id",
        "cid",
        "เลขบัตรประชาชน",
        "เลขที่บัตรประชาชน",
        "เลขประจำตัวประชาชน",
        "เลขบัตรประจำตัวประชาชน",
        "เลขประจำตัวประชาชน13หลัก",
        "เลขบัตรประชาชน13หลัก",
    },
    "first_name": {"first_name", "firstname", "ชื่อ"},
    "last_name": {"last_name", "lastname", "นามสกุล"},
    "full_name": {"full_name", "fullname", "name", "ชื่อ-นามสกุล", "ชื่อสกุล"},
    "date_of_birth": {"date_of_birth", "dob", "วันเกิด"},
    "gender": {"gender", "sex", "เพศ"},
    "province": {"province", "chw", "จังหวัด"},
    "district": {"district", "amp", "อำเภอ", "เขต"},
    "subdistrict": {"subdistrict", "tmb", "ตำบล", "แขวง"},
    "latitude": {"latitude", "lat", "ละติจูด"},
    "longitude": {"longitude", "lng", "lon", "ลองจิจูด"},
    "v1": {"v1", "smi-v1", "smiv1"},
    "v2": {"v2", "smi-v2", "smiv2"},
    "v3": {"v3", "smi-v3", "smiv3"},
    "v4": {"v4", "smi-v4", "smiv4"},
}


def _canonical(header: object) -> str:
    normalized = str(header or "").replace("\ufeff", "").strip().lower()
    compact = re.sub(r"[\s_\-./()]+", "", normalized)
    for canonical, aliases in ALIASES.items():
        if any(compact == re.sub(r"[\s_\-./()]+", "", alias) for alias in aliases):
            return canonical
    if "ประชาชน" in compact and ("เลข" in compact or "บัตร" in compact):
        return "national_id"
    return normalized


def _find_headers(rows: Any, limit: int = 20) -> tuple[list[str], int]:
    for row_number, values in enumerate(rows, start=1):
        headers = [_canonical(value) for value in values]
        if "national_id" in headers:
            return headers, row_number
        if row_number >= limit:
            break
    raise ValueError(
        "Missing national_id/เลขบัตรประชาชน column in the first 20 rows"
    )


def _national_id_text(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "")


def _json_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _parse_excel(content: bytes) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], int]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    headers, header_row = _find_headers(rows)
    parsed: dict[str, dict[str, Any]] = {}
    errors: list[dict[str, Any]] = []
    total_rows = 0
    for row_number, values in enumerate(rows, start=header_row + 1):
        if not any(value not in (None, "") for value in values):
            continue
        total_rows += 1
        raw: dict[str, Any] = dict(zip(headers, values, strict=False))
        try:
            known = {key: raw.get(key) for key in ALIASES}
            extras = {key: value for key, value in raw.items() if key not in ALIASES}
            first_name = str(known["first_name"] or "").strip()
            last_name = str(known["last_name"] or "").strip()
            full_name = str(known["full_name"] or "").strip()
            if full_name and (not first_name or not last_name):
                name_parts = full_name.split()
                first_name = first_name or name_parts[0]
                last_name = last_name or " ".join(name_parts[1:])
            if not first_name or not last_name:
                raise ValueError("first_name and last_name are required")
            raw_national_id = _national_id_text(raw.get("national_id")).strip()
            digits = re.sub(r"\D", "", raw_national_id)
            national_id_valid = len(digits) == 13
            if national_id_valid:
                digits = normalize_national_id(digits)
                national_hash = hash_national_id(digits)
            else:
                fingerprint = f"{row_number}:{raw_national_id}:{first_name}:{last_name}"
                national_hash = sha256(f"invalid:{fingerprint}".encode()).hexdigest()
                errors.append({
                    "row": row_number,
                    "message": f"เลขบัตรมี {len(digits)} หลัก (ควรมี 13 หลัก) แต่นำเข้าข้อมูลแล้ว",
                    "stored": True,
                    "raw_data": {str(key): _json_value(value) for key, value in raw.items()},
                })
            if national_hash in parsed:
                raise ValueError("Duplicate national ID in this file")
            parsed[national_hash] = {
                "national_id_hash": national_hash,
                "national_id_last4": digits[-4:] if digits else "",
                "national_id_valid": national_id_valid,
                "national_id_invalid_value": None if national_id_valid else raw_national_id,
                "first_name": first_name,
                "last_name": last_name,
                "date_of_birth": parse_date(known["date_of_birth"]),
                "gender": str(known["gender"] or "").strip() or None,
                "province": str(known["province"] or "").strip() or None,
                "district": str(known["district"] or "").strip() or None,
                "subdistrict": str(known["subdistrict"] or "").strip() or None,
                "latitude": float(known["latitude"])
                if known["latitude"] not in (None, "")
                else None,
                "longitude": float(known["longitude"])
                if known["longitude"] not in (None, "")
                else None,
                "v1": parse_bool(known["v1"]),
                "v2": parse_bool(known["v2"]),
                "v3": parse_bool(known["v3"]),
                "v4": parse_bool(known["v4"]),
                "extra_data": extras,
            }
        except (TypeError, ValueError) as exc:
            errors.append({
                "row": row_number,
                "message": str(exc),
                "raw_data": {str(key): _json_value(value) for key, value in raw.items()},
            })
    return parsed, errors, total_rows


async def preview_excel(session: AsyncSession, content: bytes) -> dict[str, Any]:
    parsed, errors, total_rows = _parse_excel(content)
    existing = set(
        await session.scalars(
            select(Patient.national_id_hash).where(
                Patient.national_id_hash.in_(list(parsed))
            )
        )
    ) if parsed else set()
    return {
        "total_rows": total_rows,
        "valid_rows": len(parsed),
        "created_count": len(set(parsed) - existing),
        "updated_count": len(existing),
        "skipped_count": len(errors),
        "errors": errors[:100],
    }


async def import_excel(
    session: AsyncSession, content: bytes, filename: str, user: User
) -> ImportJob:
    job = ImportJob(filename=filename, uploaded_by=user.id, status="processing")
    session.add(job)
    await session.flush()
    parsed, errors, _ = _parse_excel(content)
    stored_warning_rows = [int(error["row"]) for error in errors if error.get("stored")]
    if stored_warning_rows:
        previous_issues = await session.scalars(
            select(ImportIssue).where(
                ImportIssue.resolved.is_(False),
                ImportIssue.row_number.in_(stored_warning_rows),
            )
        )
        for issue in previous_issues:
            issue.resolved = True
    for error in errors:
        if error.get("stored"):
            continue
        session.add(ImportIssue(
            import_job_id=job.id,
            row_number=int(error["row"]),
            reason=str(error["message"]),
            raw_data=dict(error.get("raw_data") or {}),
        ))
    existing_patients = list(
        await session.scalars(
            select(Patient).where(Patient.national_id_hash.in_(list(parsed)))
        )
    ) if parsed else []
    existing = {patient.national_id_hash: patient for patient in existing_patients}
    for national_hash, values in parsed.items():
        patient = existing.get(national_hash)
        if patient is None:
            session.add(Patient(**values))
            job.created_count += 1
        else:
            session.add(make_version(patient, user.id))
            for key, value in values.items():
                if key not in {"national_id_hash", "national_id_last4"}:
                    setattr(patient, key, value)
            patient.version += 1
            job.updated_count += 1
    job.skipped_count = len(errors)
    job.errors = errors[:100]
    job.status = "completed_with_errors" if job.errors else "completed"
    add_audit(
        session,
        actor_id=user.id,
        action="excel.import",
        resource_type="import_job",
        resource_id=str(job.id),
        details={
            "filename": filename,
            "created": job.created_count,
            "updated": job.updated_count,
            "skipped": job.skipped_count,
        },
    )
    await session.commit()
    await session.refresh(job)
    return job
